#!/usr/bin/python
# -*- coding: utf-8 -*-
from collections import defaultdict
import time
import io
import threading
import random
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox
from os.path import exists
import requests
from fake_useragent import UserAgent
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, NamedStyle
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select

delay_choices = [8, 5, 10, 6, 20, 11]
err = False


class App(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title('Toranoana xlsx tool')
        self.face_load = tk.Frame(self)
        self.progressbar = ttk.Progressbar(
            self.face_load, length=280, mode="indeterminate")
        self.progressbar.grid(column=0, row=0, padx=10, pady=10)

        self.face_ipt = tk.Frame(self, width=300)
        self.face_ipt.pack()
        email_label = tk.Label(self.face_ipt, text='Account:')
        password_label = tk.Label(self.face_ipt, text='Password:')
        path_label = tk.Label(self.face_ipt, text='Path')
        self.email = tk.Entry(self.face_ipt)
        self.password = tk.Entry(self.face_ipt, show="*")
        self.path = tk.Entry(self.face_ipt)
        sort_button = tk.Button(self.face_ipt, text='Sort', command=self.sort)
        start_button = tk.Button(
            self.face_ipt, text='Start', command=self.start)
        email_label.grid(column=0, row=0, padx=10, pady=10)
        self.email.grid(column=1, row=0, padx=10, pady=10)
        sort_button.grid(column=2, row=0, padx=10, pady=10)
        password_label.grid(column=0, row=1, padx=10, pady=10)
        self.password.grid(column=1, row=1, padx=10, pady=10)
        start_button.grid(column=2, row=1, padx=10, pady=10)
        path_label.grid(column=0, row=2, padx=10, pady=10)
        self.path.grid(column=1, row=2, padx=10, pady=10)

    def start(self):
        email = self.email.get()
        password = self.password.get()
        if not email or not password:
            messagebox.showwarning(
                "Alert",  "Please enter email and password!")
            return
        path = self.path.get()
        if path and path[-5:] != ".xlsx":
            path += '.xlsx'
        self.face_ipt.pack_forget()
        self.face_load.pack()
        self.progressbar.start()
        t = ToranoanaXlsxTool(email, password, path)
        thread = threading.Thread(target=t.main)
        thread.start()
        self.monitor(thread)

    def sort(self):
        path = self.path.get()
        if path and path[-5:] != ".xlsx":
            path += '.xlsx'
        self.face_ipt.pack_forget()
        self.face_load.pack()
        self.progressbar.start()
        t = ToranoanaXlsxTool("", "", path)
        thread = threading.Thread(target=t.sort)
        thread.start()
        self.monitor(thread)

    def monitor(self, thread):
        if thread.is_alive():
            self.after(100, lambda: self.monitor(thread))
        else:
            self.progressbar.stop()
            self.face_load.pack_forget()
            self.face_ipt.pack()
            if err:
                messagebox.showerror('Error', err)
            else:
                messagebox.showinfo("Message",  "Done!")


class ToranoanaXlsxTool:
    def __init__(self, email, password, path):
        self.path = path if path else 'toranoana.xlsx'
        self.wb = load_workbook(self.path) if exists(self.path) else Workbook()
        self.timestamp = datetime(1, 1, 1, 0, 0)
        if exists('timestamp'):
            f = open('timestamp', 'r')
            self.timestamp = datetime.strptime(f.read(), '%Y-%m-%d')
        if 'date' not in self.wb.named_styles:
            self.wb.add_named_style(NamedStyle(
                name='date', number_format='YYYY/MM/DD'))
        self.email = email
        self.password = password

    def wait(self):
        time.sleep(random.choice(delay_choices))

    def getSheet(self, genre):
        if genre in self.wb:
            return self.wb[genre]
        else:
            sheet = self.wb.create_sheet(genre)
            titles = ["封面", "CP", "社團", "作者", "書名",
                      "價錢（日幣）", "價錢（台幣）", "出版日", "網址"]
            sheet.column_dimensions["A"].width = 100 / 7
            sheet.column_dimensions["F"].width = 16
            sheet.column_dimensions["G"].width = 16
            sheet.column_dimensions["H"].width = 16
            sheet.column_dimensions["I"].width = 55
            for i, title in enumerate(titles):
                sheet.cell(row=1, column=i+1).value = title
                sheet.cell(
                    row=1, column=i+1).alignment = Alignment(horizontal='center', vertical='center')
            return sheet

    def resizeImage(self, image):
        scale = image.height / image.width
        image.width = 100
        image.height = 100 * scale

    def getRowIndex(self, sheet, circle, author, title):
        for row in sheet.iter_rows(min_col=3, max_col=8, values_only=True):
            if ((not circle and not author) or (row[0] == circle and row[1] == author)) and row[2] == title:
                return 0
        return sheet.max_row + 1

    def getImageRow(self, image):
        if type(image.anchor) == str:
            return int(image.anchor[1:])
        else:
            return image.anchor._from.row+1

    def sort(self):
        for sheet in self.wb:
            values = []
            for row in sheet.values:
                values.append(list(row))
            values = values[1:]
            sheet_images = sheet._images
            for image in sheet_images:
                index = self.getImageRow(image)
                values[index-2][0] = image

            if sheet.title == 'unknown':
                for i, value in enumerate(values):
                    if value[0]:
                        self.resizeImage(value[0])
                        sheet.row_dimensions[i+2].height = value[0].height * 3 / 4
            else:
                circles = defaultdict(list)
                for arr in values:
                    if type(arr[7]) == str:
                        arr[7] = datetime.strptime(arr[7], "%Y/%m/%d")
                    circles[arr[2]].append(arr)
                for key in circles:
                    circles[key].sort(key=lambda x: x[7]
                                      if x[7] else datetime.now())
                sort_arr = [(key, len(circles[key])) for key in circles]
                sort_arr.sort(key=lambda x: (x[1], x[0] if x[0] else ''), reverse=True)
                row = 2
                for circle, count in sort_arr:
                    for value in circles[circle]:
                        if value[0]:
                            value[0].anchor = 'A'+str(row)
                            self.resizeImage(value[0])
                            sheet.row_dimensions[row].height = value[0].height * 3 / 4
                        sheet.cell(row=row, column=9).hyperlink = value[8]
                        for i in range(1, len(value)):
                            sheet.cell(row=row, column=i+1).value = value[i]
                        row += 1
        self.wb.save(self.path)

    def login(self):
        url = 'https://ecs.toranoana.jp/ec/app/common/login/'
        self.driver.get(url)
        self.wait()
        email = self.driver.find_element(By.ID, 'email')
        email.send_keys(self.email)
        password = self.driver.find_element(By.ID, 'repass')
        password.send_keys(self.password)
        button = self.driver.find_element(By.ID, 'submitLoginButton')
        button.click()

    def download(self):
        url = 'https://ecs.toranoana.jp/ec/app/mypage/order_history/'
        user_agent = UserAgent()
        response = self.session.get(
            url, headers={'User-Agent': user_agent.random})
        soup = BeautifulSoup(response.text, "html.parser")
        pager = soup.select_one("#pager")
        if not pager:
            return
        max_page = int(pager.get("data-maxpage"))

        for page in range(1, max_page+1):
            url = 'https://ecs.toranoana.jp/ec/app/mypage/order_history/?&currentPage=' + \
                str(page)
            response = self.session.get(
                url, headers={'User-Agent': user_agent.random})
            soup = BeautifulSoup(response.text, "html.parser")
            orders = soup.select(".hist-table4")

            for n in orders:
                order_date = datetime.strptime(n.select_one('.hist-table4-information-data').select(
                    '.hist-table4-information-data-pair')[2].select_one('.hist-table4-information-data-value').get_text(), '：%Y/%m/%d')
                if order_date < self.timestamp:
                    self.wb.save(self.path)
                    return
                url = n.select_one(
                    ".hist-table4-information-title a").get("href")
                res = requests.get(
                    url, headers={'User-Agent': user_agent.random})

                genre = image_url = couple = circle = author = title = price = date = ""
                if res.status_code == 200:
                    soup = BeautifulSoup(res.text, "html.parser")
                    if not soup.select_one(".sub-circle .sub-p span"):
                        continue
                    image_url = soup.select_one(
                        ".product-detail-image-main img").get("src")
                    spec_table = soup.select(".product-detail-spec-table tr")
                    for subsouop in spec_table:
                        t = subsouop.select_one("td").get_text()
                        if "ジャンル" in t:
                            genre = subsouop.select("td")[1].select_one(
                                "a span").get_text().replace("/", " ")
                        elif t == "カップリング":
                            couple = subsouop.select(
                                "td")[1].select_one("a span").get_text()
                        elif "発行日" in t:
                            date = datetime.strptime(subsouop.select(
                                "td")[1].select_one("a span").get_text(), "%Y/%m/%d")
                    title = soup.select_one(
                        ".product-detail-desc-title").get_text()
                    circle = soup.select_one(
                        ".sub-circle .sub-p span").get_text()
                    author = soup.select_one(".sub-name .sub-p a").get_text()
                    price = int(soup.select_one(".pricearea__price").get_text().replace(
                        "円 （税込） ", "").replace(",", ""))
                else:
                    genre = 'unknown'
                    image_url = n.select_one(
                        '.hist-table4-information-thumbnail img').get('src')
                    title = n.select_one(
                        '.hist-table4-information-title a span').get_text()
                    price = int(n.select('.hist-table4-information-data')[1].select('.hist-table4-information-data-pair')[
                        2].select_one('.hist-table4-information-data-value').get_text()[1:-1].replace(",", ""))
                image_res = requests.get(image_url)
                image_file = io.BytesIO(image_res.content)
                image = Image(image_file)
                self.resizeImage(image)

                print(title)
                sheet = self.getSheet(genre)
                r = self.getRowIndex(sheet, circle, author, title)
                if r == 0:
                    self.wait()
                    continue
                values = [couple, circle, author, title, price, "", date, url]

                sheet.add_image(image, 'A' + str(r))
                for row in sheet.iter_rows(min_row=r, min_col=2, max_row=r, max_col=len(values)+1):
                    for i, cell in enumerate(row):
                        cell.value = values[i]
                        if i == 6:
                            cell.style = 'date'
                        if i == 7:
                            cell.hyperlink = values[i]
                            cell.style = "Hyperlink"
                        if i < 4:
                            sheet.column_dimensions[cell.column_letter].width = max(
                                sheet.column_dimensions[cell.column_letter].width, len(values[i])*2.5)
                        cell.alignment = Alignment(
                            horizontal='center', vertical='center')
                self.wait()
        if "Sheet" in self.wb:
            del self.wb["Sheet"]
        self.wb.save(self.path)

    def main(self):
        self.driver = webdriver.Chrome()
        self.driver.minimize_window()
        self.login()
        self.wait()
        cookies = self.driver.get_cookies()
        self.session = requests.Session()
        for cookie in cookies:
            self.session.cookies.set(cookie['name'], cookie['value'])
        self.session.headers.update({
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
            "Accept-Encoding": "gzip, deflate, br",
            "Accept-Language": "zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7,ja;q=0.6,zh-CN;q=0.5,ig;q=0.4,el;q=0.3",
            "Cache-Control": "max-age=0",
            "Sec-Ch-Ua": "\" Not A;Brand\";v=\"99\", \"Chromium\";v=\"101\", \"Google Chrome\";v=\"101\"",
            "Sec-Ch-Ua-Mobile": "?0",
            "Sec-Ch-Ua-Platform": "\"Windows\"",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "cross-site",
            "Sec-Fetch-User": "?1",
            "Sec-Gpc": "1",
            "Upgrade-Insecure-Requests": "1",
        })
        self.driver.get(
            'https://ecs.toranoana.jp/ec/app/mypage/order_history/')
        ship_list = ['not_shipped', 'shipped']
        period_list = ['this_year', 'last_year', 'two_years_ago']

        for s in ship_list:
            try:
                radio = self.driver.find_element(
                    By.ID, s).find_element(By.XPATH, '..')
            except:
                global err
                err = 'Login Error'
            radio.click()
            for p in period_list:
                select = Select(self.driver.find_element(
                    By.ID, 'searchPeriod'))
                button = self.driver.find_element(By.ID, 'submit')
                select.select_by_value(p)
                button.click()
                self.wait()
                self.download()
        self.sort()
        timestamp = open("timestamp", "w")
        timestamp.write(datetime.now().strftime('%Y-%m-%d'))
        timestamp.close()


app = App()
app.mainloop()
